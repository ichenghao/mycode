'''
本程序实现从管道壁厚计算书中把不同等级不同尺寸管道壁厚找出来并写入一个Excel文件中
1. 使用了openpyxl库和fractions库
2. 原始管道壁厚计算书文件名：thick_cal.xlsx；原始文件以等级命名工作表，工作表内容格式固定
3. 最后生成的文件名：pipe_thick.xlsx；尺寸按从小到大排序，等级按照原工作表排序
本程序在Python 3.7.3下运行成功
'''

# 导入openpyxl库
import openpyxl as xl

# 加载管道壁厚计算书
wb1=xl.load_workbook('thick_cal.xlsx')

# 根据各个工作表名称找出所有等级，存放在spec列表里
spec=[x for x in wb1.sheetnames if (len(x)==3 and x!='TOC')]

# 遍历各等级工作表中C列，找出所有的尺寸，存放在size列表里
size=[]
for x in spec:
    ws=wb1[x]
    for i in range(10,50): #这里找尺寸时最多到第49行
        cell='C'+str(i)
        if ws[cell].value==None:
            break
        size.append(ws[cell].value)
size=list(set(size)) #尺寸进行去重操作

# size从小到大排序
import fractions #加载fractions库
size_float=[] #新建一个列表用于存储转换为浮点数的尺寸
for x in size:
    if ' ' in x: #如果有空格说明含有整数部分和分数部分
        y=x.strip('"').split()
        z=float(y[0])+fractions.Fraction(y[1])
    else:
        y=x.strip('"')
        z=0.0+fractions.Fraction(y)
    size_float.append(z)
d=dict(zip(size_float,size))   #将两个列表打包成字典，浮点数为key，字符串为value
size=[d[i] for i in sorted(d)] #按照key从小到大排序形成新的value列表

# 打印所有等级和尺寸
print(f"等级数量为{len(spec)},包括：{spec}")
print(f"尺寸数量为{len(size)},包括：{size}")

# 新建一个工作簿ws2，将spec和size分别写入第一行和第一列
wb2=xl.Workbook()
ws2=wb2.active
x=['']
x.extend(spec)
ws2.append(x)
row=2
for i in size:
    ws2.cell(column=1,row=row).value=i
    row+=1

# 找出行和列的边界
row_limit=len(size)+2
col_limit=len(spec)+2

# 从ws1中找到壁厚，写入ws2相应的单元格中
for row in range(2,row_limit):
    for col in range(2,col_limit):
        spec=ws2.cell(column=col, row=1).value
        size=ws2.cell(column=1,row=row).value
        ws1=wb1[spec]
        for i in range(10,50): #这里找尺寸时最多到第49行
            j=ws1.cell(column=3,row=i).value
            if j==None:
                break
            elif j==size:
                ws2.cell(column=col,row=row).value=ws1.cell(column=7,row=i).value

# 将wb2保存名为pipe_thick.xlsx的文件
wb2.save('pipe_thick.xlsx')

# 运行成功，打印信息
print('管道壁厚提取成功!')
