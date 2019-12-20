''' by Cheng Hao
本程序实现了HAZOP报告中提取信息输出到output.xlsx
1. 原数据文件名称：hazop.xlsx
2. 模板文件：sheets.xlsx
3. 输出结果文件：output.xlsx
本程序在Python 3.7.3下运行成功
'''

# 为了计算程序运行时间导入time库
import time
start_time=time.time()
print("Start running. It may take several minutes.")

# 导入openpyxl库
import openpyxl as xl

# 加载两个文件
wb1=xl.load_workbook('hazop.xlsx', read_only=True)
ws1=wb1.active
wb2=xl.load_workbook('sheets.xlsx')

# 打印加载文件用的时间
print(f"Original files loaded, load time: {(time.time()-start_time)/60:.1f} mins.")

# 定义函数处理P&ID No
def pidno(s):
    if s != None:
        a=s.split('-')
        if len(a)==3:
            b='HFY-3800-'+a[0]+'-PRO-PID-'+a[2]
            return b
        else:
            return s

# 定义函数合并Recommendation相同的项
def combine(ws,ws2):
    for i in ['B8','A18','B6','B7','B9','A12','A14','A16','B4']:
        if ws2[i].value != None and ws[i].value != ws2[i].value:
            if ws[i].value != None:
                ws[i].value+='\n'+ws2[i].value
            else:
                ws[i].value=ws2[i].value

# 开始执行
print("Processing...", end='')

# 定义原文件的最大行边界
row_limit=3500

for row in range(2,row_limit):
    recommendation=ws1.cell(column=14,row=row).value
    if recommendation != None:
        title='R'+recommendation.split('.',1)[0].zfill(3)
        print('.',end='')
        if title not in wb2.sheetnames:
            ws=wb2.copy_worksheet(wb2['temp'])
            ws.title=title                                                  # sheet title
            ws['B8'].value=title[1:]                                            # recommendation no
            ws['A18'].value=recommendation.split('.',1)[1]                      # recommendation
            ws['B6'].value=ws1.cell(column=1,row=row).value                     # node
            ws['B7'].value=ws1.cell(column=2,row=row).value                     # topic
            ws['B9'].value=ws1.cell(column=15,row=row).value.split('.',1)[1]    # response by
            ws['A12'].value=ws1.cell(column=6,row=row).value                    # item
            ws['A14'].value=ws1.cell(column=7,row=row).value                    # cause
            ws['A16'].value=ws1.cell(column=8,row=row).value                    # consequence
            ws['B4'].value=pidno(ws1.cell(column=16,row=row).value)             # pid no
        else:
            ws=wb2[title]
            # 放到ws2里暂存，然后合并到ws中，最后把ws2删除
            ws2=wb2.copy_worksheet(wb2['temp'])
            ws2['B8'].value=title[1:]                                            # recommendation no
            ws2['A18'].value=recommendation.split('.',1)[1]                      # recommendation
            ws2['B6'].value=ws1.cell(column=1,row=row).value                     # node
            ws2['B7'].value=ws1.cell(column=2,row=row).value                     # topic
            ws2['B9'].value=ws1.cell(column=15,row=row).value.split('.',1)[1]    # response by
            ws2['A12'].value=ws1.cell(column=6,row=row).value                    # item
            ws2['A14'].value=ws1.cell(column=7,row=row).value                    # cause
            ws2['A16'].value=ws1.cell(column=8,row=row).value                    # consequence
            ws2['B4'].value=pidno(ws1.cell(column=16,row=row).value)             # pid no
            combine(ws,ws2)
            wb2.remove(ws2)

# 删除工作簿中的模板工作表
wb2.remove(wb2['temp'])

# 按工作表名称排序
#wb2._sheets.sort(key=lambda ws:ws.title)

# 保存成果文件
wb2.save('output.xlsx')

# 打印程序执行时间
print(f"\nFinish, running time: {(time.time()-start_time)/60:.1f} mins.")
